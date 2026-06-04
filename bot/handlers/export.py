import io
import logging
from datetime import date

import openpyxl
from aiogram import Router
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, Message

from bot import database

logger = logging.getLogger(__name__)
router = Router(name="export")


@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    if message.from_user is None:
        return

    user_id = message.from_user.id
    stats = await database.repository.get_user_stats(user_id)

    if not stats or stats.get("total_messages", 0) == 0:
        await message.answer("\U0001f50d Ҳали овозли хабар йўқ. Экспорт учун аввал овозли хабар юборинг!")
        return

    await message.answer("\U0001f4ca Excel файл тайёрланмоқда\u2026")

    # Барча хабарларни олиш
    first = stats.get("first_message")
    if first:
        start_date = first.date() if hasattr(first, 'date') else date.fromisoformat(str(first)[:10])
    else:
        start_date = date.today()

    # Workbook яратиш
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транскрипциялар"

    # Сарлавҳалар
    headers = ["№", "Сана", "Вақт", "Давомийлик (сек)", "Транскрипция"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # Маълумотлар — охирги 30 кун
    from datetime import timedelta
    all_messages = []
    current = date.today()
    for i in range(30):
        d = current - timedelta(days=i)
        msgs = await database.repository.get_voice_messages_by_date(user_id, d)
        all_messages.extend(msgs)

    # Сатрлар
    for row_num, msg in enumerate(all_messages, 2):
        created = msg.get("created_at")
        if hasattr(created, 'date'):
            date_str = created.strftime("%d.%m.%Y")
            time_str = created.strftime("%H:%M")
        else:
            created_str = str(created) or ""
            date_str = created_str[:10]
            time_str = created_str[11:16]

        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=date_str)
        ws.cell(row=row_num, column=3, value=time_str)
        ws.cell(row=row_num, column=4, value=msg.get("duration") or 0)
        ws.cell(row=row_num, column=5, value=msg.get("transcript") or "")

    # Устун кенглиги
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60

    # Файлни юбориш
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ovozxotira_{date.today().strftime('%Y%m%d')}.xlsx"
    await message.answer_document(
        BufferedInputFile(buffer.getvalue(), filename=filename),
        caption=f"\U0001f4ca <b>Транскрипциялар экспорти</b>\n"
                f"\U0001f4dd Жами: {stats['total_messages']} та хабар\n"
                f"\U0001f4c5 Охирги 30 кун"
    )
    logger.info("Export sent to user_id=%d", user_id)